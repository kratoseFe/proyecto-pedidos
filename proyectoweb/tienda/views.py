from django.shortcuts import render
from django.views.generic import TemplateView, View
from tienda.models import Producto
from openpyxl import Workbook
from django.http.response import HttpResponse
from tienda.utils import render_pdf

def tienda (request):
    product = Producto.objects.all()
    return render(request, "tienda.html", {'productos':product,})

def reporte (request):
    product = Producto.objects.all()
    return render(request, "listado.html", {'productos':product,})

class ReporteProductoExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        producto = Producto.objects.all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de productos"
        ws['A1'] = "Reporte productos"
        ws.merge_cells('A1:C1')

        ws['A3'] = 'Id'
        ws['B3'] = 'Producto'
        ws['C3'] = 'Precio'
        ws['D3'] = 'Imagen'
        ws['E3'] = 'created'
        cont = 4

        for product in producto:
            ws.cell(row=cont, column=1).value = product.id
            ws.cell(row=cont, column=2).value = product.nombre
            ws.cell(row=cont, column=3).value = product.precio
            ws.cell(row=cont, column=4).value = product. imagen.url
            ws.cell(row=cont, column=5).value = product.created
            cont += 1

        nombre_archivo = "ReportesProductosExcel.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['content-Disposition'] = content
        wb.save(response)
        return response
    
def listado (request):
    product = Producto.objects.all()
    return render(request, "listado.html", {'productos':product,})  

class ListaProductoPdf(View):
    def get(self, request, *args, **kwargs):
        productos = Producto.objects.all()
        data = {
            "productos":productos
        }
        pdf = render_pdf('tienda/listado.html', data)
        return HttpResponse(pdf, content_type='application/pdf')